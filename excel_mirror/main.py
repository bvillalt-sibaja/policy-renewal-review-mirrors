"""Excel mirror: a small Tkinter app that mimics genuine Excel navigation (a Name
Box + formula bar, exactly like real Excel) over a real .xlsx workbook, so an RPA
automation drives it the same way a human would - type a cell reference into the
Name Box and press Enter to select it, then type a value into the formula bar and
press Enter to commit - rather than manipulating the file invisibly via openpyxl
with no visible app at all.

Usage: main.py <path-to-xlsx> [sheet-name]
Every commit is saved to the workbook on disk immediately (openpyxl load+save per
edit - fine for the small dummy workbooks this mirrors).
"""
import re
import sys

import tkinter as tk
from tkinter import ttk, font as tkfont

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("excel_mirror: openpyxl is required (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)

GRID_ROWS = 25
GRID_COLS = 12
HEADER_BG = "#F3F3F3"
GRID_BG = "#FFFFFF"
SELECTED_BG = "#E8F2EC"
BORDER = "#D0D0D0"

# Real Excel brand colors: dark-green ribbon accent, light-gray column/row headers,
# a slightly darker green for the active/selected column & row header.
EXCEL_GREEN = "#185C37"
EXCEL_GREEN_ACCENT = "#217346"
RIBBON_TAB_BG = "#F3F2F1"
COLHEADER_BG = "#F3F3F3"
COLHEADER_SELECTED_BG = "#CAEAD8"
GRIDLINE = "#D4D4D4"

CELL_REF_RE = re.compile(r"^([A-Za-z]+)([0-9]+)$")


def cell_ref_to_rc(ref):
    m = CELL_REF_RE.match(ref.strip())
    if not m:
        return None
    col_letters, row_str = m.groups()
    return int(row_str), column_index_from_string(col_letters.upper())


class ExcelMirror:
    def __init__(self, root, path, sheet_name=None):
        self.root = root
        self.path = path
        self.sheet_name = sheet_name
        self.selected = (1, 1)

        root.title(f"{path.split('/')[-1]} - Excel Mirror")
        # Fixed size AND position every launch - same reasoning as the Outlook
        # mirror's own SELECTORS.md: a deterministic window position/geometry is
        # what makes coordinate-based automation (Name Box, formula bar) reliable.
        # NOTE: changing anything about the layout below (ribbon height, toolbar
        # height, row height) shifts the absolute screen coordinates the
        # automation clicks (Excel Mirror Go To Cell in renewal_review.robot) -
        # recalibrate those live after any layout change here.
        root.geometry("1000x600+100+100")
        root.configure(bg=GRID_BG)

        mono = self._pick_font(root, "Calibri", "Segoe UI", "Helvetica", size=11)
        self.mono = mono
        tab_font = self._pick_font(root, "Segoe UI", "Calibri", "Helvetica", size=10)
        title_font = self._pick_font(root, "Segoe UI Semibold", "Calibri", "Helvetica", size=11)

        # -- title bar strip (dark Excel green, with the app "icon" and file name) --
        titlebar = tk.Frame(root, bg=EXCEL_GREEN, height=30)
        titlebar.pack(fill="x", side="top")
        titlebar.pack_propagate(False)
        tk.Label(
            titlebar, text=f"⬚  {path.split('/')[-1]}", font=title_font,
            fg="#FFFFFF", bg=EXCEL_GREEN, anchor="w",
        ).pack(side="left", padx=10)

        # -- decorative ribbon tab strip (not functional - visual authenticity only) --
        ribbon = tk.Frame(root, bg=RIBBON_TAB_BG, height=28, bd=0, highlightthickness=1,
                           highlightbackground=BORDER)
        ribbon.pack(fill="x", side="top")
        ribbon.pack_propagate(False)
        for i, tab in enumerate(["File", "Home", "Insert", "Page Layout", "Formulas", "Data", "Review", "View"]):
            fg = "#FFFFFF" if tab == "Home" else "#3B3A39"
            bg = EXCEL_GREEN_ACCENT if tab == "Home" else RIBBON_TAB_BG
            tk.Label(
                ribbon, text=tab, font=tab_font, fg=fg, bg=bg, padx=10, pady=5,
            ).pack(side="left")

        # -- formula bar row: Name Box + fx + formula bar, styled like real Excel --
        toolbar = tk.Frame(root, bg="#FFFFFF", height=30, bd=0, highlightthickness=1,
                            highlightbackground=BORDER)
        toolbar.pack(fill="x", side="top")
        toolbar.pack_propagate(False)

        self.name_box = tk.Entry(toolbar, width=12, font=mono, relief="solid", bd=1,
                                  highlightthickness=0)
        self.name_box.pack(side="left", padx=(6, 4), pady=4)
        self.name_box.bind("<Return>", self._on_name_box_enter)

        tk.Frame(toolbar, bg=BORDER, width=1).pack(side="left", fill="y", pady=4)
        tk.Label(toolbar, text="fx", bg="#FFFFFF", fg=EXCEL_GREEN, font=(mono.actual("family"), 11, "italic"),
                 width=3).pack(side="left", padx=(6, 2))

        self.formula_bar = tk.Entry(toolbar, font=mono, relief="solid", bd=1, highlightthickness=0)
        self.formula_bar.pack(side="left", fill="x", expand=True, padx=(0, 6), pady=4)
        self.formula_bar.bind("<Return>", self._on_formula_bar_enter)

        grid_frame = tk.Frame(root, bg=GRID_BG)
        grid_frame.pack(fill="both", expand=True)

        columns = ["#"] + [get_column_letter(c) for c in range(1, GRID_COLS + 1)]
        self.tree = ttk.Treeview(grid_frame, columns=columns, show="headings", selectmode="browse", height=GRID_ROWS)
        style = ttk.Style()
        style.theme_use(style.theme_use())  # keep current theme, just override colors below
        style.configure(
            "Treeview", font=mono, rowheight=22, background=GRID_BG,
            fieldbackground=GRID_BG, bordercolor=GRIDLINE, borderwidth=1,
        )
        style.configure(
            "Treeview.Heading", font=(mono.actual("family"), 10), background=COLHEADER_BG,
            foreground="#3B3A39", relief="ridge", borderwidth=1,
        )
        style.map(
            "Treeview",
            background=[("selected", SELECTED_BG)],
            foreground=[("selected", "#000000")],
        )
        style.map("Treeview.Heading", background=[("active", COLHEADER_SELECTED_BG)])
        self.tree.heading("#", text="")
        self.tree.column("#", width=36, anchor="center", stretch=False)
        for c in range(1, GRID_COLS + 1):
            letter = get_column_letter(c)
            self.tree.heading(letter, text=letter)
            self.tree.column(letter, width=110, anchor="w", stretch=False)
        self.tree.tag_configure("rowlabel", background=COLHEADER_BG, foreground="#3B3A39")
        self.tree.pack(fill="both", expand=True, side="left")

        vsb = ttk.Scrollbar(grid_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        # -- status bar, like real Excel's bottom strip --
        statusbar = tk.Frame(root, bg=EXCEL_GREEN_ACCENT, height=22)
        statusbar.pack(fill="x", side="bottom")
        statusbar.pack_propagate(False)
        tk.Label(statusbar, text="Ready", font=tab_font, fg="#FFFFFF",
                 bg=EXCEL_GREEN_ACCENT).pack(side="left", padx=10)

        self._load_workbook()
        self._render_grid()
        self._select_cell(1, 1)

    def _pick_font(self, root, *candidates, size=11):
        available = set(tkfont.families(root))
        for name in candidates:
            if name in available:
                return tkfont.Font(family=name, size=size)
        return tkfont.Font(size=size)

    def _load_workbook(self):
        self.wb = load_workbook(self.path)
        self.ws = self.wb[self.sheet_name] if self.sheet_name else self.wb.active

    def _render_grid(self):
        self.tree.delete(*self.tree.get_children())
        for r in range(1, GRID_ROWS + 1):
            values = []
            for c in range(1, GRID_COLS + 1):
                cell = self.ws.cell(row=r, column=c)
                values.append("" if cell.value is None else str(cell.value))
            self.tree.insert("", "end", iid=str(r), values=[str(r)] + values)

    def _select_cell(self, row, col):
        self.selected = (row, col)
        ref = f"{get_column_letter(col)}{row}"
        self.name_box.delete(0, "end")
        self.name_box.insert(0, ref)
        value = self.ws.cell(row=row, column=col).value
        self.formula_bar.delete(0, "end")
        self.formula_bar.insert(0, "" if value is None else str(value))
        try:
            self.tree.selection_set(str(row))
            self.tree.see(str(row))
        except tk.TclError:
            pass

    def _on_tree_select(self, _event):
        sel = self.tree.selection()
        if not sel:
            return
        row = int(sel[0])
        if row == self.selected[0]:
            return  # avoid the selection_set() in _select_cell re-triggering this
        self._select_cell(row, self.selected[1])

    def _on_name_box_enter(self, _event):
        ref = self.name_box.get()
        rc = cell_ref_to_rc(ref)
        if rc is None:
            return
        row, col = rc
        if row < 1 or col < 1 or row > GRID_ROWS or col > GRID_COLS:
            return
        self._select_cell(row, col)
        self.formula_bar.focus_set()

    def _on_formula_bar_enter(self, _event):
        row, col = self.selected
        value = self.formula_bar.get()
        self.ws.cell(row=row, column=col, value=value)
        self.wb.save(self.path)
        self._load_workbook()
        self._render_grid()
        self._select_cell(row, col)
        self.name_box.focus_set()


def main(path, sheet_name=None):
    root = tk.Tk()
    ExcelMirror(root, path, sheet_name)
    root.mainloop()


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print("usage: main.py <path-to-xlsx> [sheet-name]", file=sys.stderr)
        sys.exit(2)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) == 3 else None)
